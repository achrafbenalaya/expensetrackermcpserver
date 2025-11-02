import os
import io
import re
from datetime import datetime, date
from typing import Optional, Dict, Any, List
import csv

from openpyxl import load_workbook
from openpyxl import Workbook
from azure.storage.blob import ContainerClient
from azure.core.exceptions import ResourceNotFoundError


def _parse_yyyymm(value: str) -> str:
    if re.fullmatch(r"\d{4}-\d{2}", value):
        return value
    if re.fullmatch(r"\d{6}", value):
        return f"{value[:4]}-{value[4:]}"
    raise ValueError("month must be 'YYYY-MM' or 'YYYYMM'")


def _parse_date(value: str) -> date:
    return datetime.strptime(value, "%Y-%m-%d").date()


class ExpenseTracker:
    """
    Lightweight storage backed by Azure Blob Storage using a SAS container URL.
    Uses csv and openpyxl instead of pandas for smaller dependency footprint.
    """

    def __init__(
        self,
        container_sas_url: Optional[str] = None,
        blob_name: Optional[str] = None,
        default_monthly_budget: Optional[float] = None,
    ):
        self.container_sas_url = (
            container_sas_url
            or os.environ.get("EXPENSES_CONTAINER_SAS_URL")
            or os.environ.get("BUDGETTRACKER_CONTAINER_SAS_URL")
        )
        if not self.container_sas_url:
            raise RuntimeError(
                "Missing SAS container URL. Set EXPENSES_CONTAINER_SAS_URL or BUDGETTRACKER_CONTAINER_SAS_URL."
            )

        self.blob_name = blob_name or os.environ.get("EXPENSES_BLOB_NAME") or "expenses.xlsx"
        self.default_monthly_budget = (
            default_monthly_budget
            if default_monthly_budget is not None
            else float(os.environ.get("EXPENSE_BUDGET_DEFAULT", "1000"))
        )

        self.container = ContainerClient.from_container_url(self.container_sas_url)

    # ---------- Blob <-> rows helpers ----------

    def _empty_rows(self) -> List[Dict[str, Any]]:
        return []

    def _download_rows(self) -> List[Dict[str, Any]]:
        blob = self.container.get_blob_client(self.blob_name)
        try:
            data = blob.download_blob().readall()
            if self.blob_name.lower().endswith(".xlsx"):
                wb = load_workbook(filename=io.BytesIO(data), read_only=True, data_only=True)
                ws = wb.active
                rows = list(ws.iter_rows(values_only=True))
                if not rows:
                    return self._empty_rows()
                headers = [str(h) for h in rows[0]]
                result = []
                for row in rows[1:]:
                    item = {headers[i]: row[i] for i in range(len(headers))}
                    if item.get("Date") is None:
                        item["Date"] = ""
                    if item.get("Amount") is None:
                        item["Amount"] = 0.0
                    if item.get("Category") is None:
                        item["Category"] = ""
                    if item.get("Description") is None:
                        item["Description"] = ""
                    result.append(item)
                return result
            else:
                text = data.decode("utf-8")
                reader = csv.DictReader(io.StringIO(text))
                return [r for r in reader]
        except ResourceNotFoundError:
            return self._empty_rows()

    def _upload_rows(self, rows: List[Dict[str, Any]]) -> None:
        blob = self.container.get_blob_client(self.blob_name)
        if self.blob_name.lower().endswith(".xlsx"):
            wb = Workbook()
            ws = wb.active
            headers = ["Date", "Amount", "Category", "Description"]
            ws.append(headers)
            for r in rows:
                ws.append([r.get("Date", ""), r.get("Amount", ""), r.get("Category", ""), r.get("Description", "")])
            buf = io.BytesIO()
            wb.save(buf)
            buf.seek(0)
            blob.upload_blob(buf, overwrite=True)
        else:
            headers = ["Date", "Amount", "Category", "Description"]
            sio = io.StringIO()
            writer = csv.DictWriter(sio, fieldnames=headers)
            writer.writeheader()
            for r in rows:
                writer.writerow({
                    "Date": r.get("Date", ""),
                    "Amount": r.get("Amount", ""),
                    "Category": r.get("Category", ""),
                    "Description": r.get("Description", ""),
                })
            blob.upload_blob(sio.getvalue().encode("utf-8"), overwrite=True)

    # ---------- Public API (the MCP tools) ----------

    def get_expenses(self, month: str, category: Optional[str] = None) -> List[Dict[str, Any]]:
        month = _parse_yyyymm(month)
        rows = self._download_rows()
        result = []
        for r in rows:
            d = r.get("Date")
            if not d:
                continue
            if isinstance(d, datetime):
                ddate = d.date()
            elif isinstance(d, date):
                ddate = d
            else:
                try:
                    ddate = datetime.strptime(str(d), "%Y-%m-%d").date()
                except Exception:
                    continue
            row_month = f"{ddate.year:04d}-{ddate.month:02d}"
            if row_month != month:
                continue
            cat = r.get("Category") or ""
            if category and str(cat).strip().lower() != category.strip().lower():
                continue
            amount = r.get("Amount", 0.0)
            try:
                amount = float(amount)
            except Exception:
                amount = 0.0
            result.append({
                "Date": ddate.isoformat(),
                "Amount": amount,
                "Category": str(cat),
                "Description": str(r.get("Description", "")),
            })
        return result

    def add_expense(self, amount: float, category: str, description: str, date_str: str) -> Dict[str, Any]:
        dt = _parse_date(date_str)
        try:
            amount = float(amount)
        except Exception:
            raise ValueError("amount must be numeric")
        rows = self._download_rows()
        new_row = {"Date": dt.isoformat(), "Amount": amount, "Category": category, "Description": description}
        rows.append(new_row)
        self._upload_rows(rows)
        return new_row

    def get_budget_status(self, month: str) -> Dict[str, Any]:
        month = _parse_yyyymm(month)
        budget_env_key = f"EXPENSE_BUDGET_{month.replace('-', '')}"
        budget = float(os.environ.get(budget_env_key, self.default_monthly_budget))
        expenses = self.get_expenses(month=month, category=None)
        total_spent = sum(e.get("Amount", 0.0) for e in expenses)
        remaining = budget - total_spent
        return {
            "month": month,
            "budget": budget,
            "total_expenses": round(total_spent, 2),
            "remaining": round(remaining, 2),
            "currency": os.environ.get("EXPENSE_CURRENCY", "USD"),
        }

    def get_spending_summary(self, start_date: str, end_date: str) -> Dict[str, Any]:
        start_dt = _parse_date(start_date)
        end_dt = _parse_date(end_date)
        if start_dt > end_dt:
            raise ValueError("start_date must be on or before end_date")
        rows = self._download_rows()
        total = 0.0
        by_cat: Dict[str, float] = {}
        for r in rows:
            d = r.get("Date")
            if not d:
                continue
            if isinstance(d, datetime):
                ddate = d.date()
            elif isinstance(d, date):
                ddate = d
            else:
                try:
                    ddate = datetime.strptime(str(d), "%Y-%m-%d").date()
                except Exception:
                    continue
            if ddate < start_dt or ddate > end_dt:
                continue
            amt = r.get("Amount", 0.0)
            try:
                amt = float(amt)
            except Exception:
                amt = 0.0
            total += amt
            cat = r.get("Category") or "Uncategorized"
            by_cat[str(cat)] = by_cat.get(str(cat), 0.0) + amt
        by_cat = {k: round(float(v), 2) for k, v in by_cat.items()}
        return {"from": start_dt.isoformat(), "to": end_dt.isoformat(), "total": round(total, 2), "by_category": by_cat}

    def categorize_expense(self, description: str) -> Dict[str, Any]:
        desc = (description or "").lower()
        rules = {
            "groceries": ["market", "grocery", "supermarket", "aldi", "lidl", "costco"],
            "transport": ["uber", "taxi", "metro", "bus", "train", "gas", "fuel"],
            "restaurants": ["restaurant", "diner", "cafe", "coffee", "bar", "takeaway"],
            "utilities": ["electric", "water", "gas bill", "internet", "phone"],
            "shopping": ["amazon", "mall", "store", "retail", "clothes", "electronics"],
            "health": ["pharmacy", "doctor", "clinic", "dentist", "hospital"],
            "entertainment": ["netflix", "cinema", "movie", "music", "concert", "game"],
            "travel": ["hotel", "flight", "airline", "airbnb", "booking.com"],
            "education": ["course", "tuition", "book", "udemy", "coursera"],
            "other": [],
        }
        for cat, keywords in rules.items():
            for kw in keywords:
                if kw in desc:
                    return {"category": cat, "matched": kw}
        return {"category": "other", "matched": None}
